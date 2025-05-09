import matplotlib.pyplot as plt
import openpyxl
from matplotlib.pyplot import figure

import json

# # Write Data into Excel file
# data = openpyxl.load_workbook('data-1.xlsx')
# sheet = data.active  # data.get_sheet_by_name('empsheet')
#
# for r in range(1, 60):
#     for c in range(1, 4):
#         sheet.cell(row=r, column=c).value = "ABU BAKAKR"
#
# data.save('data-1.xlsx')
# ******************************************************************************************************
# data = openpyxl.load_workbook('data-1.xlsx')
# sheet = data.active  # data.get_sheet_by_name('empsheet')
#
# # This script for Read all dat from perticular excel sheeet
# row = sheet.max_row
# col = sheet.max_column
#
#
# for r in range(1, row + 1):
#     for c in range(1, col + 1):
#         print(sheet.cell(row=r, column=c).value, end="         ")
#     print('')
# *****************************************************************************************************
# def getRowCount(file, sheetName):
#     data = openpyxl.load_workbook(file)
#     sheet = data.get_sheet_by_name(sheetName)
#     return sheet.max_row
#
#
# def getColumnCount(file, sheetName):
#     data = openpyxl.load_workbook(file)
#     sheet = data.get_sheet_by_name(sheetName)
#     return sheet.max_column
#
#
# def readData(file, sheetName, rowno, columnno):
#     data = openpyxl.load_workbook(file)
#     sheet = data.get_sheet_by_name(sheetName)
#     return sheet.cell(row=rowno, column=columnno).value
#
#
# def writeData(file, sheetName, rowno, columnno, data):
#     data = openpyxl.load_workbook(file)
#     sheet = data.get_sheet_by_name(sheetName)
#     sheet.cell(row=rowno, column=columnno).value = data
#     data.save(file)
#
#
# row = getRowCount('data-1.xlsx', 'Sheet1')
# col = getColumnCount('data-1.xlsx', 'Sheet1')
# data = readData('data-1.xlsx', 'Sheet1', 3, 1)
# print(row, col, data)

data = json.load(open('stock.json'))
data = data['stock_name']
new_data = {}
for k, v in data.items():
    new_data[k] = len(v)

print(new_data)
figure(num=None, figsize=(16, 6), dpi=80, facecolor='w', edgecolor='k')
plt.axhline(y=0, color='k')
plt.bar(new_data.keys(), new_data.values())
plt.show()
